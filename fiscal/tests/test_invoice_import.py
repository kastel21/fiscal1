"""Tests for Invoice Excel import. Invoice 01 rules."""

from decimal import Decimal
from io import BytesIO

from django.test import TestCase

from openpyxl import Workbook

from fiscal.models import FiscalDevice
from fiscal.services.excel_parser import (
    list_and_rank_sheets,
    parse_excel,
    validate_line_math,
)
from fiscal.services.invoice_import_service import validate_invoice_import


class ExcelParserInvoice01Tests(TestCase):
    def test_header_detection_invoice01(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Invoice 01"
        ws.append([None] * 3)
        ws.append(["Qty", "Description", "Amount Due", "Total"])
        ws.append([2, "Product A", 25.0, 50.0])
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        lines, meta = parse_excel(buf.read(), sheet_name="Invoice 01")
        self.assertEqual(len(lines), 1)
        self.assertEqual(lines[0]["quantity"], 2)
        self.assertEqual(lines[0]["line_total"], 50.0)
        self.assertIn("header_row", meta)

    def test_empty_row_skipping(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["Qty", "Description", "Total"])
        ws.append([1, "A", 10])
        ws.append([None, None, None])
        ws.append([2, "B", 20])
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        lines, _ = parse_excel(buf.read())
        self.assertEqual(len(lines), 2)

    def test_unit_price_derivation(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["Qty", "Description", "Total"])
        ws.append([2, "Item", 100])
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        lines, _ = parse_excel(buf.read())
        self.assertEqual(len(lines), 1)
        self.assertEqual(lines[0]["unit_price"], 50.0)

    def test_validate_line_math_mismatch(self):
        line = {"row_num": 5, "quantity": 2, "unit_price": 10, "line_total": 30, "description": "X"}
        errors = validate_line_math(line)
        self.assertTrue(any("does not equal" in e for e in errors))

    def test_validate_line_math_ok(self):
        line = {"row_num": 5, "quantity": 2, "unit_price": 10, "line_total": 20, "description": "X"}
        errors = validate_line_math(line)
        self.assertEqual(len(errors), 0)


class SheetRankingTests(TestCase):
    def test_invoice01_preferred(self):
        wb = Workbook()
        ws1 = wb.create_sheet("quote", 0)
        ws2 = wb.create_sheet("Invoice 01", 1)
        ws3 = wb.create_sheet("Delivery Note 1", 2)
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        sheets = list_and_rank_sheets(buf.read())
        names = [s["name"] for s in sheets]
        self.assertIn("Invoice 01", names)
        inv01 = next(s for s in sheets if s["name"] == "Invoice 01")
        self.assertTrue(inv01["importable"])
        quote = next(s for s in sheets if s["name"] == "quote")
        self.assertFalse(quote["importable"])


class ValidateInvoiceImportTests(TestCase):
    def setUp(self):
        self.device = FiscalDevice.objects.create(
            device_id=99999,
            certificate_pem="x",
            private_key_pem="x",
            is_registered=True,
        )

    def test_missing_tax_blocked(self):
        errors = validate_invoice_import(
            [{"quantity": 1, "description": "X", "unit_price": 10, "line_total": 10}],
            "FiscalInvoice",
            "USD",
            None,
            self.device,
        )
        self.assertTrue(any("tax" in e.lower() for e in errors))
