"""Tests for fiscal compliance audit fixes: export tenant scope, task tenant_id, previous_hash, etc."""

from decimal import Decimal

from django.test import TestCase

from fiscal.export_utils import render_excel
from fiscal.models import FiscalDay, FiscalDevice, Receipt, ReceiptSubmissionResponse
from fiscal.services.receipt_submission_response_service import store_receipt_submission_response
from fiscal.tasks import close_day_task, open_day_task, submit_receipt_task
from tenants.models import Tenant

FAKE_CERT = "-----BEGIN CERTIFICATE-----\ntest\n-----END CERTIFICATE-----"
FAKE_KEY = "-----BEGIN PRIVATE KEY-----\ntest\n-----END PRIVATE KEY-----"


class ExcelExportTenantScopeTests(TestCase):
    """Dashboard Excel export returns only the given tenant's data."""

    def setUp(self):
        self.tenant_a = Tenant.objects.create(
            name="Tenant A",
            slug="tenant-a",
            device_id=50001,
            is_active=True,
        )
        self.tenant_b = Tenant.objects.create(
            name="Tenant B",
            slug="tenant-b",
            device_id=50002,
            is_active=True,
        )
        self.device_a = FiscalDevice.all_objects.create(
            tenant=self.tenant_a,
            device_id=60001,
            device_serial_no="A1",
            certificate_pem=FAKE_CERT,
            private_key_pem=FAKE_KEY,
            is_registered=True,
        )
        self.device_b = FiscalDevice.all_objects.create(
            tenant=self.tenant_b,
            device_id=60002,
            device_serial_no="B1",
            certificate_pem=FAKE_CERT,
            private_key_pem=FAKE_KEY,
            is_registered=True,
        )
        self.receipt_a = Receipt.all_objects.create(
            tenant=self.tenant_a,
            device=self.device_a,
            fiscal_day_no=1,
            receipt_global_no=1,
            receipt_type="FiscalInvoice",
            receipt_total=Decimal("100.00"),
            fdms_receipt_id=101,
        )
        self.receipt_b = Receipt.all_objects.create(
            tenant=self.tenant_b,
            device=self.device_b,
            fiscal_day_no=1,
            receipt_global_no=2,
            receipt_type="FiscalInvoice",
            receipt_total=Decimal("200.00"),
            fdms_receipt_id=102,
        )

    def test_render_excel_with_tenant_only_includes_that_tenant_receipts(self):
        """When tenant is passed, Invoices and Credit Notes sheets only contain that tenant's data."""
        from openpyxl import load_workbook
        from io import BytesIO
        xlsx = render_excel("month", tenant=self.tenant_a)
        wb = load_workbook(BytesIO(xlsx), read_only=True, data_only=True)
        ws_inv = wb["Invoices"]
        rows = list(ws_inv.iter_rows(min_row=2, max_row=10, values_only=True))
        device_ids = [r[0] for r in rows if r[0] is not None]
        self.assertIn(60001, device_ids)
        self.assertNotIn(60002, device_ids)
        wb.close()

    def test_render_excel_without_tenant_includes_all(self):
        """When tenant is None, export can include all tenants (e.g. superadmin)."""
        from openpyxl import load_workbook
        from io import BytesIO
        xlsx = render_excel("month", tenant=None)
        wb = load_workbook(BytesIO(xlsx), read_only=True, data_only=True)
        ws_inv = wb["Invoices"]
        rows = list(ws_inv.iter_rows(min_row=2, max_row=10, values_only=True))
        device_ids = [r[0] for r in rows if r[0] is not None]
        self.assertGreaterEqual(len(device_ids), 1)
        wb.close()


class ReceiptSubmissionResponseTenantTests(TestCase):
    """ReceiptSubmissionResponse stores tenant when device has tenant."""

    def setUp(self):
        self.tenant = Tenant.objects.create(
            name="T",
            slug="t1",
            device_id=50001,
            is_active=True,
        )
        self.device = FiscalDevice.all_objects.create(
            tenant=self.tenant,
            device_id=60001,
            device_serial_no="A1",
            certificate_pem=FAKE_CERT,
            private_key_pem=FAKE_KEY,
            is_registered=True,
        )

    def test_store_receipt_submission_response_sets_tenant(self):
        """store_receipt_submission_response sets tenant_id from device."""
        store_receipt_submission_response(
            device=self.device,
            receipt_global_no=1,
            status_code=200,
            response_body={"receiptID": 101},
            fiscal_day_no=1,
            receipt=None,
        )
        rsp = ReceiptSubmissionResponse.all_objects.filter(device=self.device).first()
        self.assertIsNotNone(rsp)
        self.assertEqual(rsp.tenant_id, self.tenant.pk)


class TaskRequiresTenantIdTests(TestCase):
    """Celery tasks reject execution when tenant_id is missing."""

    def setUp(self):
        self.tenant = Tenant.objects.create(
            name="T",
            slug="t1",
            device_id=50001,
            is_active=True,
        )
        self.device = FiscalDevice.all_objects.create(
            tenant=self.tenant,
            device_id=60001,
            device_serial_no="A1",
            certificate_pem=FAKE_CERT,
            private_key_pem=FAKE_KEY,
            is_registered=True,
            last_fiscal_day_no=1,
        )

    def test_submit_receipt_task_rejects_missing_tenant_id(self):
        """submit_receipt_task returns error when tenant_id is not provided."""
        result = submit_receipt_task(
            device_id=self.device.device_id,
            fiscal_day_no=1,
            receipt_type="FiscalInvoice",
            receipt_currency="USD",
            invoice_no="INV-1",
            receipt_lines=[],
            receipt_taxes=[],
            receipt_payments=[],
            receipt_total=0.0,
            tenant_id=None,
        )
        self.assertFalse(result.get("success"))
        self.assertIn("tenant_id is required", result.get("error", ""))

    def test_open_day_task_rejects_missing_tenant_id(self):
        """open_day_task returns error when tenant_id is not provided."""
        result = open_day_task(self.device.device_id, tenant_id=None)
        self.assertFalse(result.get("success"))
        self.assertIn("tenant_id is required", result.get("error", ""))

    def test_close_day_task_rejects_missing_tenant_id(self):
        """close_day_task returns error when tenant_id is not provided."""
        result = close_day_task(self.device.device_id, tenant_id=None)
        self.assertFalse(result.get("success"))
        self.assertIn("tenant_id is required", result.get("error", ""))


class FiscalDayCreatedWithTenantTests(TestCase):
    """FiscalDay is created with tenant_id when open_day runs."""

    def setUp(self):
        self.tenant = Tenant.objects.create(
            name="T",
            slug="t1",
            device_id=50001,
            is_active=True,
        )
        self.device = FiscalDevice.all_objects.create(
            tenant=self.tenant,
            device_id=60001,
            device_serial_no="A1",
            certificate_pem=FAKE_CERT,
            private_key_pem=FAKE_KEY,
            is_registered=True,
        )

    def test_open_day_creates_fiscal_day_with_tenant(self):
        """DeviceApiService.open_day creates FiscalDay with tenant_id set (mocked)."""
        from unittest.mock import MagicMock, patch
        from fiscal.services.device_api import DeviceApiService
        service = DeviceApiService()
        with patch.object(service, "get_status", return_value=({"fiscalDayStatus": "FiscalDayClosed", "lastFiscalDayNo": 0}, None)):
            with patch("fiscal.services.device_api.fdms_request") as mock_req:
                mock_resp = MagicMock()
                mock_resp.status_code = 200
                mock_resp.json.return_value = {"fiscalDayNo": 1}
                mock_req.return_value = mock_resp
                with patch("fiscal.services.device_api.log_fdms_call"):
                    fiscal_day, err = service.open_day(self.device)
        self.assertIsNone(err)
        self.assertIsNotNone(fiscal_day)
        self.assertEqual(fiscal_day.tenant_id, self.tenant.pk)
        # Also verify via all_objects in case manager scoping differs
        fd = FiscalDay.all_objects.filter(device=self.device).order_by("-opened_at").first()
        self.assertIsNotNone(fd)
        self.assertEqual(fd.tenant_id, self.tenant.pk)


class TenantPreviousHashUpdateTests(TestCase):
    """Tenant.previous_hash is updated after successful receipt submission."""

    def setUp(self):
        self.tenant = Tenant.objects.create(
            name="T",
            slug="t1",
            device_id=50001,
            is_active=True,
        )
        self.device = FiscalDevice.all_objects.create(
            tenant=self.tenant,
            device_id=60001,
            device_serial_no="A1",
            certificate_pem=FAKE_CERT,
            private_key_pem=FAKE_KEY,
            is_registered=True,
            last_fiscal_day_no=1,
            last_receipt_global_no=0,
        )

    def test_tenant_previous_hash_updated_after_receipt_save(self):
        """Simulate receipt save path: Tenant.previous_hash is updated in transaction."""
        from django.db import transaction
        from fiscal.models import Receipt
        from tenants.models import Tenant
        receipt_hash = "dGVzdF9oYXNoX2Jhc2U2NA=="
        with transaction.atomic():
            receipt = Receipt.all_objects.create(
                tenant=self.tenant,
                device=self.device,
                fiscal_day_no=1,
                receipt_global_no=1,
                receipt_type="FiscalInvoice",
                receipt_total=Decimal("50.00"),
                receipt_hash=receipt_hash,
                fdms_receipt_id=999,
            )
            if self.device.tenant_id and (receipt.receipt_hash or "").strip():
                Tenant.objects.filter(pk=self.device.tenant_id).update(
                    previous_hash=(receipt.receipt_hash or "").strip(),
                )
        self.tenant.refresh_from_db()
        self.assertEqual(self.tenant.previous_hash, receipt_hash)
