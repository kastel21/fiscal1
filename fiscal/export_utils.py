"""Audit-ready PDF and Excel exports. Read-only, immutable."""

from datetime import datetime
from io import BytesIO

from django.utils import timezone


def render_pdf(data: dict, range_key: str) -> bytes:
    """Generate PDF report. Uses reportlab if available, else minimal text PDF."""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
        from reportlab.lib import colors
    except ImportError:
        return _pdf_fallback(data, range_key)

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    story.append(Paragraph("FDMS Dashboard Export", styles["Title"]))
    story.append(Spacer(1, 12))
    story.append(Paragraph(f"Period: {range_key} | Generated: {timezone.now().isoformat()}", styles["Normal"]))
    story.append(Spacer(1, 12))

    status = data.get("status", {})
    story.append(Paragraph("Status", styles["Heading2"]))
    story.append(Paragraph(f"Fiscal Day: {status.get('fiscalDay', 'N/A')}", styles["Normal"]))
    story.append(Paragraph(f"FDMS Connectivity: {status.get('fdmsConnectivity', 'N/A')}", styles["Normal"]))
    story.append(Paragraph(f"Certificate: {status.get('certificate', 'N/A')}", styles["Normal"]))
    story.append(Paragraph(f"Last Sync: {status.get('lastSync', 'N/A')}", styles["Normal"]))
    story.append(Spacer(1, 12))

    metrics = data.get("metrics", {})
    story.append(Paragraph("Metrics", styles["Heading2"]))
    story.append(Paragraph(f"Invoices fiscalised: {metrics.get('invoicesFiscalised', 0)}", styles["Normal"]))
    story.append(Paragraph(f"Credit notes: {metrics.get('creditNotes', 0)}", styles["Normal"]))
    story.append(Paragraph(f"Net total: {metrics.get('netTotal', 0)}", styles["Normal"]))
    story.append(Paragraph(f"VAT total: {metrics.get('vatTotal', 0)}", styles["Normal"]))
    story.append(Spacer(1, 12))

    comp = data.get("compliance", {})
    story.append(Paragraph("Compliance", styles["Heading2"]))
    story.append(Paragraph(f"Last receiptGlobalNo: {comp.get('lastReceiptGlobalNo', 'N/A')}", styles["Normal"]))
    story.append(Paragraph(f"Last OpenDay: {comp.get('lastOpenDay') or 'N/A'}", styles["Normal"]))
    story.append(Paragraph(f"Last CloseDay: {comp.get('lastCloseDay') or 'N/A'}", styles["Normal"]))
    story.append(Paragraph(f"Last Ping: {comp.get('lastPing') or 'N/A'}", styles["Normal"]))
    if comp.get('reportingFrequency') is not None:
        story.append(Paragraph(f"Reporting frequency: {comp.get('reportingFrequency')} min", styles["Normal"]))

    doc.build(story)
    return buf.getvalue()


def _pdf_fallback(data: dict, range_key: str) -> bytes:
    """Minimal text-based output when reportlab not installed."""
    lines = [
        "FDMS Dashboard Export",
        f"Period: {range_key}",
        f"Generated: {timezone.now().isoformat()}",
        "",
        "Status:",
        f"  Fiscal Day: {data.get('status', {}).get('fiscalDay', 'N/A')}",
        f"  Certificate: {data.get('status', {}).get('certificate', 'N/A')}",
        "",
        "Metrics:",
        f"  Invoices: {data.get('metrics', {}).get('invoicesFiscalised', 0)}",
        f"  Credit notes: {data.get('metrics', {}).get('creditNotes', 0)}",
        f"  Net total: {data.get('metrics', {}).get('netTotal', 0)}",
    ]
    return ("\n".join(lines)).encode("utf-8")


def render_excel(range_key: str, tenant=None) -> bytes:
    """Generate Excel workbook. Requires openpyxl. When tenant is provided, all data is scoped to that tenant."""
    from openpyxl import Workbook
    from fiscal.services.dashboard_service import _date_range, get_summary, get_errors
    from fiscal.models import Receipt

    data = get_summary(None, range_key, tenant=tenant)
    start_dt, end_dt = _date_range(range_key)
    wb = Workbook()

    ws = wb.active
    ws.title = "Summary"
    ws.append(["FDMS Dashboard Export", range_key, timezone.now().isoformat()])
    ws.append([])
    ws.append(["Status", "Value"])
    for k, v in (data.get("status") or {}).items():
        ws.append([k, str(v)])
    ws.append([])
    ws.append(["Metrics", "Value"])
    for k, v in (data.get("metrics") or {}).items():
        ws.append([k, v])
    ws.append([])
    ws.append(["Compliance", "Value"])
    for k, v in (data.get("compliance") or {}).items():
        ws.append([k, str(v)])

    ws2 = wb.create_sheet("Invoices")
    invoices_qs = Receipt.all_objects.filter(
        receipt_type="FiscalInvoice",
        fdms_receipt_id__isnull=False,
    ).exclude(fdms_receipt_id=0).filter(
        created_at__gte=start_dt,
        created_at__lte=end_dt,
    )
    if tenant is not None:
        invoices_qs = invoices_qs.filter(tenant=tenant)
    invoices = list(invoices_qs.order_by("-created_at")[:500])
    ws2.append(["Device", "FiscalDay", "ReceiptGlobalNo", "InvoiceNo", "Total", "Created"])
    for r in invoices:
        ws2.append([r.device.device_id, r.fiscal_day_no, r.receipt_global_no, r.invoice_no or "", float(r.receipt_total or 0), r.created_at.isoformat() if r.created_at else ""])

    ws3 = wb.create_sheet("Credit Notes")
    cns_qs = Receipt.all_objects.filter(
        receipt_type="CreditNote",
        fdms_receipt_id__isnull=False,
    ).exclude(fdms_receipt_id=0).filter(
        created_at__gte=start_dt,
        created_at__lte=end_dt,
    )
    if tenant is not None:
        cns_qs = cns_qs.filter(tenant=tenant)
    cns = list(cns_qs.order_by("-created_at")[:500])
    ws3.append(["Device", "FiscalDay", "ReceiptGlobalNo", "InvoiceNo", "Total", "OriginalInvoice", "Created"])
    for r in cns:
        ws3.append([r.device.device_id, r.fiscal_day_no, r.receipt_global_no, r.invoice_no or "", float(r.receipt_total or 0), r.original_invoice_no or "", r.created_at.isoformat() if r.created_at else ""])

    ws4 = wb.create_sheet("Errors")
    errors = get_errors(None, range_key, tenant=tenant)
    ws4.append(["Endpoint", "Status", "Error", "Created"])
    for e in errors:
        ws4.append([e.get("endpoint", ""), e.get("statusCode"), e.get("error", ""), e.get("createdAt", "")])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def render_invoice_pdf(ctx: dict) -> bytes:
    """Generate FDMS-compliant invoice PDF. Same layout as receipt_invoice.html."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.utils import ImageReader
    from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    from reportlab.lib import colors

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []

    title = ctx.get("document_type", "TAX INVOICE")
    if title == "Fiscal Credit Note":
        title = "CREDIT NOTE"
    elif title == "Fiscal Tax Invoice":
        title = "TAX INVOICE"
    story.append(Paragraph(title, styles["Title"]))
    story.append(Spacer(1, 12))

    if ctx.get("original_invoice_no") or ctx.get("reason"):
        story.append(Paragraph("Original Invoice / Reason", styles["Heading2"]))
        if ctx.get("original_invoice_no"):
            story.append(Paragraph(f"Original Invoice Number: {ctx['original_invoice_no']}", styles["Normal"]))
        if ctx.get("original_invoice_date"):
            story.append(Paragraph(f"Original Invoice Date: {ctx['original_invoice_date']}", styles["Normal"]))
        if ctx.get("reason"):
            story.append(Paragraph(f"Reason: {ctx['reason']}", styles["Normal"]))
        story.append(Spacer(1, 8))

    story.append(Paragraph("Supplier / Device Identity", styles["Heading2"]))
    story.append(Paragraph(ctx.get("business_name", ""), styles["Normal"]))
    if ctx.get("business_address"):
        story.append(Paragraph(ctx["business_address"], styles["Normal"]))
    if ctx.get("vat_number"):
        story.append(Paragraph(f"VAT / Taxpayer No: {ctx['vat_number']}", styles["Normal"]))
    story.append(Paragraph(f"FDMS Device ID: {ctx.get('device_id', '')}", styles["Normal"]))
    story.append(Spacer(1, 8))

    story.append(Paragraph("Fiscal Header", styles["Heading2"]))
    story.append(Paragraph(f"Receipt Global No: {ctx.get('receipt_global_no', '')}", styles["Normal"]))
    story.append(Paragraph(f"Fiscalisation Date & Time: {ctx.get('fiscal_date', '')}", styles["Normal"]))
    story.append(Paragraph(f"Currency: {ctx.get('currency', '')}", styles["Normal"]))
    story.append(Spacer(1, 8))

    lines = ctx.get("lines", [])
    if lines:
        data = [["Description", "Qty", "Unit Price", "Line Total"]]
        for line in lines:
            data.append([
                str(line.get("description", ""))[:40],
                str(line.get("quantity", "")),
                "%.2f" % line.get("unit_price", 0),
                "%.2f" % line.get("line_total", 0),
            ])
        t = Table(data, colWidths=[200, 50, 80, 80])
        t.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.grey), ("GRID", (0, 0), (-1, -1), 0.5, colors.black)]))
        story.append(Paragraph("Line Items", styles["Heading2"]))
        story.append(t)
        story.append(Spacer(1, 8))

    story.append(Paragraph("Totals", styles["Heading2"]))
    curr = ctx.get("currency", "")
    totals_data = [
        ["Subtotal", "%.2f %s" % (ctx.get("subtotal", 0), curr)],
        ["Total Tax", "%.2f %s" % (ctx.get("total_tax", 0), curr)],
        ["Grand Total", "%.2f %s" % (ctx.get("grand_total", 0), curr)],
    ]
    t2 = Table(totals_data, colWidths=[120, 150])
    t2.setStyle(TableStyle([("GRID", (0, 0), (-1, -1), 0.5, colors.black), ("FONT", (0, -1), (-1, -1), "Helvetica-Bold", 10)]))
    story.append(t2)
    story.append(Spacer(1, 8))

    payments = ctx.get("payment_rows", [])
    if payments:
        pay_data = [["Payment Method", "Amount Paid"]]
        for p in payments:
            pay_data.append([str(p.get("method", "")), "%.2f %s" % (p.get("amount", 0), curr)])
        t3 = Table(pay_data, colWidths=[120, 150])
        t3.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.grey), ("GRID", (0, 0), (-1, -1), 0.5, colors.black)]))
        story.append(Paragraph("Payment Summary", styles["Heading2"]))
        story.append(t3)
        story.append(Spacer(1, 8))

    story.append(Paragraph("QR Verification", styles["Heading2"]))
    qr_code_value = (ctx.get("qr_code_value") or "").strip()
    if qr_code_value:
        try:
            import qrcode
            qr_buf = BytesIO()
            qrcode.make(qr_code_value).save(qr_buf, format="PNG")
            qr_buf.seek(0)
            qr_width_pt = 40 * 72 / 25.4
            qr_img = Image(ImageReader(qr_buf), width=qr_width_pt, height=qr_width_pt)
            story.append(qr_img)
            story.append(Spacer(1, 6))
            story.append(Paragraph(qr_code_value.replace("/", "&#47;"), styles["Normal"]))
        except Exception:
            story.append(Paragraph("Fiscalised. Verify with ZIMRA.", styles["Normal"]))
    else:
        story.append(Paragraph("Fiscalised. Verify with ZIMRA.", styles["Normal"]))

    doc.build(story)
    return buf.getvalue()


