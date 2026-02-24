"""
Reusable Excel import engine. Header auto-detection, line extraction.
Used for Invoice and Credit Note imports.
"""

import re
from decimal import Decimal
from io import BytesIO
from typing import Any

REQUIRED_HEADER_MATCHES = 3
HEADER_KEYWORDS = {
    "qty": ["qty", "quantity", "qty.", "quant"],
    "description": ["description", "item", "product", "name", "details"],
    "amount": ["amount", "unit price", "price", "rate", "amount due"],
    "total": ["total", "line total", "amount", "value", "sum"],
    "uom": ["uom", "unit"],
}
INVOICE_SHEETS_ALLOWED = ["Invoice 01", "Invoice01", "invoice 01"]
INVOICE_SHEETS_IGNORED = ["quote", "delivery note", "Delivery Note"]
ROUNDING_TOLERANCE = Decimal("0.02")
ALL_KEYWORDS = set()
for v in HEADER_KEYWORDS.values():
    ALL_KEYWORDS.update(v)


def _normalize(val: Any) -> str:
    if val is None:
        return ""
    s = str(val).strip().lower()
    return re.sub(r"\s+", " ", s)


def _score_header_row(row_values: list) -> int:
    """Return number of keyword matches. Row is header if >= REQUIRED_HEADER_MATCHES."""
    matches = 0
    combined = " ".join(_normalize(c) for c in row_values)
    for kw in ALL_KEYWORDS:
        if kw in combined:
            matches += 1
    return matches


def _detect_column_indices(header_row: list) -> dict[str, int | None]:
    """Map canonical field names to column indices. Returns {qty, description, unit_price, line_total}."""
    result = {"qty": None, "description": None, "unit_price": None, "line_total": None}
    for idx, cell in enumerate(header_row):
        val = _normalize(cell)
        if not val:
            continue
        for field, keywords in HEADER_KEYWORDS.items():
            if any(kw in val for kw in keywords):
                if field == "qty" and result["qty"] is None:
                    result["qty"] = idx
                elif field == "description" and result["description"] is None:
                    result["description"] = idx
                elif field == "amount" and result["unit_price"] is None:
                    result["unit_price"] = idx
                elif field == "total":
                    if result["line_total"] is None or "total" in val:
                        result["line_total"] = idx
                break
    return result


def _to_decimal(val: Any) -> Decimal:
    if val is None or val == "":
        return Decimal("0")
    try:
        s = str(val).replace(",", "").strip()
        return Decimal(s)
    except Exception:
        return Decimal("0")


def _extract_line(
    row: list,
    indices: dict[str, int | None],
    row_num: int,
) -> dict | None:
    """Extract one line item. Returns None if invalid."""
    def get(idx: int | None) -> Any:
        if idx is not None and 0 <= idx < len(row):
            v = row[idx]
            return v
        return None

    qty = _to_decimal(get(indices["qty"]))
    desc = str(get(indices["description"]) or "").strip()
    unit_price = _to_decimal(get(indices["unit_price"]))
    line_total = _to_decimal(get(indices["line_total"]))

    if qty <= 0 and line_total <= 0:
        return None
    if not desc and line_total <= 0:
        return None

    if line_total > 0 and qty <= 0:
        qty = Decimal("1")
    if line_total > 0 and unit_price <= 0 and qty > 0:
        unit_price = line_total / qty

    unit_price_derived = unit_price <= 0 and line_total > 0 and qty > 0
    if unit_price_derived:
        unit_price = line_total / qty
    return {
        "row_num": row_num,
        "quantity": float(qty),
        "description": desc or "(no description)",
        "unit_price": float(unit_price),
        "line_total": float(line_total),
        "unit_price_derived": unit_price_derived,
        "from_excel": True,
    }


def list_and_rank_sheets(file_content: bytes) -> list[dict]:
    """List sheets and rank by Invoice 01 likelihood. Returns [{name, score, importable}]."""
    from openpyxl import load_workbook
    wb = load_workbook(read_only=True, data_only=True, filename=BytesIO(file_content))
    result = []
    for name in wb.sheetnames:
        nlower = name.strip().lower()
        score = 0
        importable = True
        for allowed in INVOICE_SHEETS_ALLOWED:
            if allowed.lower() in nlower or nlower in allowed.lower():
                score = 10
                break
        for ign in INVOICE_SHEETS_IGNORED:
            if ign in nlower:
                importable = False
                score = -1
                break
        if importable and score == 0:
            score = 1
        result.append({"name": name, "score": score, "importable": importable})
    wb.close()
    result.sort(key=lambda x: (-x["score"], x["name"]))
    return result


def validate_line_math(line: dict) -> list[str]:
    """Return list of validation errors for a line. qty×unit_price ≈ line_total."""
    errors = []
    qty = Decimal(str(line.get("quantity", 0)))
    up = Decimal(str(line.get("unit_price", 0)))
    lt = Decimal(str(line.get("line_total", 0)))
    if qty <= 0:
        errors.append(f"Row {line.get('row_num', '?')}: Quantity must be > 0.")
    if lt <= 0:
        errors.append(f"Row {line.get('row_num', '?')}: Line total must be > 0.")
    if not (line.get("description") or "").strip():
        errors.append(f"Row {line.get('row_num', '?')}: Description is required.")
    expected = (qty * up).quantize(Decimal("0.01"))
    diff = abs(expected - lt)
    if diff > ROUNDING_TOLERANCE:
        errors.append(
            f"Row {line.get('row_num', '?')}: Quantity × Unit Price ({float(expected):.2f}) does not equal Line Total ({float(lt):.2f}). "
            "Please correct the Excel file or adjust values."
        )
    return errors


def parse_excel(file_content: bytes, sheet_name: str | None = None) -> tuple[list[dict], dict]:
    """
    Parse Excel file. Returns (lines, metadata).
    metadata: {sheet_name, header_row, detected_columns}.
    For Credit Note: treat all amounts as credit values (positive in preview).
    """
    from openpyxl import load_workbook

    wb = load_workbook(read_only=True, data_only=True, filename=BytesIO(file_content))
    ws = wb.active
    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.active
        sheet_name = ws.title

    header_row_idx = None
    header_row_values: list = []
    for i, row in enumerate(ws.iter_rows(max_row=50, values_only=True), 1):
        if not row:
            continue
        score = _score_header_row(row)
        if score >= REQUIRED_HEADER_MATCHES:
            header_row_idx = i
            header_row_values = list(row)
            break

    if header_row_idx is None:
        wb.close()
        return [], {"sheet_name": sheet_name, "header_row": None, "error": "No header row detected"}

    indices = _detect_column_indices(header_row_values)
    valid_indices = [i for i in (indices["qty"], indices["description"], indices["unit_price"], indices["line_total"]) if i is not None]
    if indices["line_total"] is None and valid_indices:
        indices["line_total"] = max(valid_indices) + 1

    lines = []
    for i, row in enumerate(ws.iter_rows(min_row=header_row_idx + 1, values_only=True), header_row_idx + 1):
        row_list = list(row) if row else []
        line = _extract_line(row_list, indices, i)
        if line and line["line_total"] > 0:
            lines.append(line)

    wb.close()
    return lines, {
        "sheet_name": sheet_name,
        "header_row": header_row_idx,
        "detected_columns": {k: v for k, v in indices.items() if v is not None},
    }
